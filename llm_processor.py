#!/usr/bin/env python3
"""
LLM 处理器 - 用于自动提取作者单位和生成中文总结
使用 OpenAI 兼容 API 格式
"""

import os
import json
import time
import requests
from pathlib import Path
from datetime import date

# ==================== API 配置 ====================
API_BASE_URL = os.environ.get("LLM_API_BASE", "https://uni-api.cstcloud.cn/v1")
API_KEY = os.environ.get("LLM_API_KEY", "")
MODEL_NAME = os.environ.get("LLM_MODEL", "gpt-4o")  # 可配置模型名称

# ==================== PDF 处理 ====================
def extract_pdf_text(pdf_path: str, max_pages: int = 3) -> str:
    """从 PDF 提取文本（仅前几页）"""
    try:
        import fitz  # PyMuPDF
        
        if not Path(pdf_path).exists():
            print(f"[WARN] PDF not found: {pdf_path}")
            return ""
        
        doc = fitz.open(pdf_path)
        text_parts = []
        
        for page_num in range(min(max_pages, len(doc))):
            page = doc[page_num]
            text_parts.append(page.get_text())
        
        doc.close()
        return "\n".join(text_parts)
    
    except Exception as e:
        print(f"[ERROR] PDF extraction failed: {e}")
        return ""


def call_llm_api(prompt: str, system_prompt: str = "") -> str:
    """调用 LLM API"""
    if not API_KEY:
        raise Exception("LLM_API_KEY not configured")
    
    url = f"{API_BASE_URL}/chat/completions"
    
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})
    
    payload = {
        "model": MODEL_NAME,
        "messages": messages,
        "temperature": 0.1,
        "max_tokens": 500
    }
    
    response = requests.post(url, headers=headers, json=payload, timeout=60)
    response.raise_for_status()
    
    result = response.json()
    content = result["choices"][0]["message"]["content"]
    return content.strip()


def extract_affiliations(pdf_text: str, authors: list) -> str:
    """从 PDF 文本中提取作者单位"""
    if not pdf_text:
        return ""
    
    system_prompt = """你是一个学术信息提取助手。请从论文中提取作者单位信息。
只返回单位名称，用分号分隔。不要包含其他说明文字。
如果没有找到单位信息，返回空字符串。"""
    
    authors_str = ", ".join(authors) if isinstance(authors, list) else str(authors)
    
    prompt = f"""
论文作者：{authors_str}

论文文本（前几页）:
{pdf_text[:8000]}  # 限制长度避免超时

请提取所有作者的单位信息，返回格式：
单位 1; 单位 2; ...

如果没有找到单位信息，返回空字符串。
"""
    
    try:
        result = call_llm_api(prompt, system_prompt)
        # 清理结果
        result = result.replace("单位：", "").replace("作者单位：", "").strip()
        return result
    except Exception as e:
        print(f"[ERROR] Affiliation extraction failed: {e}")
        return ""


def generate_summary_cn(abstract: str) -> str:
    """基于摘要生成中文总结"""
    if not abstract:
        return ""
    
    system_prompt = """你是一个学术论文摘要翻译助手。
请将英文摘要翻译成简洁的中文总结，150 字以内。
只返回中文总结，不要包含其他说明文字。"""
    
    prompt = f"""
请将以下论文摘要翻译成简洁的中文总结（150 字以内）:

{abstract}

中文总结:
"""
    
    try:
        result = call_llm_api(prompt, system_prompt)
        # 清理结果
        result = result.replace("中文总结：", "").replace("总结：", "").strip()
        # 确保长度限制
        if len(result) > 150:
            result = result[:147] + "..."
        return result
    except Exception as e:
        print(f"[ERROR] Summary generation failed: {e}")
        return ""


# ==================== 主流程 ====================
def process_paper(paper: dict, papers_dir: Path) -> dict:
    """处理单篇论文"""
    arxiv_id = paper.get("arxiv_id", "")
    title = paper.get("title", "")
    authors = paper.get("authors", [])
    abstract = paper.get("abstract", "")
    
    print(f"\n[LLM] Processing: {arxiv_id} - {title[:50]}...")
    
    # 1. 提取 PDF 文本
    pdf_path = paper.get("pdf_local_path", "")
    if not pdf_path:
        pdf_path = str(papers_dir / f"{arxiv_id}.pdf")
    
    pdf_text = extract_pdf_text(pdf_path, max_pages=2)
    
    # 2. 提取作者单位
    print(f"[LLM] Extracting affiliations...")
    affiliations = extract_affiliations(pdf_text, authors)
    print(f"[LLM] Affiliations: {affiliations[:100]}..." if affiliations else "[LLM] No affiliations found")
    
    # 3. 生成中文总结
    print(f"[LLM] Generating Chinese summary...")
    summary_cn = generate_summary_cn(abstract)
    print(f"[LLM] Summary: {summary_cn[:50]}..." if summary_cn else "[LLM] No summary generated")
    
    # 4. 延迟避免 API 限流
    time.sleep(1)
    
    return {
        "arxiv_id": arxiv_id,
        "affiliations": affiliations,
        "summary_cn": summary_cn,
        "success": bool(affiliations or summary_cn)
    }


def load_pending_papers() -> list:
    """加载待处理的论文"""
    output_json = Path("new_papers.json")
    if not output_json.exists():
        return []
    
    data = json.loads(output_json.read_text(encoding="utf-8"))
    return data.get("papers_to_process", [])


def update_excel_with_results(results: list):
    """将 LLM 结果更新到 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment
        
        excel_path = Path("papers_record.xlsx")
        if not excel_path.exists():
            print(f"[WARN] Excel not found: {excel_path}")
            return False
        
        wb = openpyxl.load_workbook(str(excel_path))
        if "Papers" not in wb.sheetnames:
            print("[WARN] Papers sheet not found")
            return False
        
        ws = wb["Papers"]
        
        # 获取列索引
        header_row = [cell.value for cell in ws[1]]
        arxiv_col = header_row.index("arxiv_id") + 1 if "arxiv_id" in header_row else 0
        affiliations_col = header_row.index("affiliations") + 1 if "affiliations" in header_row else 0
        summary_cn_col = header_row.index("summary_cn") + 1 if "summary_cn" in header_row else 0
        
        if not all([arxiv_col, affiliations_col, summary_cn_col]):
            print("[ERROR] Required columns not found")
            return False
        
        # 更新每一行
        updated = 0
        for row in range(2, ws.max_row + 1):
            arxiv_id = ws.cell(row=row, column=arxiv_col).value
            if not arxiv_id:
                continue
            
            for result in results:
                if result["arxiv_id"] == arxiv_id:
                    if result["affiliations"]:
                        ws.cell(row=row, column=affiliations_col, value=result["affiliations"])
                    if result["summary_cn"]:
                        ws.cell(row=row, column=summary_cn_col, value=result["summary_cn"])
                    
                    # 设置单元格格式
                    for col in [affiliations_col, summary_cn_col]:
                        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
                    
                    updated += 1
                    print(f"[INFO] Updated: {arxiv_id}")
                    break
        
        if updated > 0:
            wb.save(str(excel_path))
            print(f"[INFO] Excel updated: {updated} records")
            return True
        else:
            print("[WARN] No records updated")
            return False
            
    except Exception as e:
        print(f"[ERROR] Excel update failed: {e}")
        return False


def export_viewer_json():
    """导出 viewer/papers_data.json"""
    try:
        import openpyxl
        
        excel_path = Path("papers_record.xlsx")
        viewer_json = Path("viewer/papers_data.json")
        
        if not excel_path.exists():
            return False
        
        wb = openpyxl.load_workbook(str(excel_path), read_only=True)
        if "Papers" not in wb.sheetnames:
            return False
        
        ws = wb["Papers"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return False
        
        headers = [str(h) if h is not None else "" for h in header_row]
        index = {name: i for i, name in enumerate(headers)}
        
        required = [
            "arxiv_id", "title", "authors", "affiliations", "published_date",
            "categories", "abstract", "summary_cn", "pdf_filename", "crawled_date", "notes"
        ]
        missing = [c for c in required if c not in index]
        if missing:
            print(f"[WARN] Missing columns: {missing}")
            return False
        
        def norm(v):
            if v is None:
                return ""
            return str(v).replace("\n", " ").strip()
        
        papers_by_id = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            paper = {col: norm(row[index[col]]) for col in required}
            if not paper["arxiv_id"]:
                continue
            paper["pdf_url"] = f"https://arxiv.org/pdf/{paper['arxiv_id']}"
            
            # 质量优先：有中文总结和单位的优先
            if paper["summary_cn"] or paper["affiliations"]:
                arxiv_id = paper["arxiv_id"]
                if arxiv_id not in papers_by_id:
                    papers_by_id[arxiv_id] = paper
                else:
                    old = papers_by_id[arxiv_id]
                    # 保留信息更完整的
                    if (len(paper.get("summary_cn", "")) + len(paper.get("affiliations", "")) >
                        len(old.get("summary_cn", "")) + len(old.get("affiliations", ""))):
                        papers_by_id[arxiv_id] = paper
        
        papers = list(papers_by_id.values())
        papers.sort(key=lambda x: (x["crawled_date"], x["published_date"], x["arxiv_id"]), reverse=True)
        
        crawled_dates = sorted({p["crawled_date"] for p in papers if p["crawled_date"]})
        published_dates = sorted({p["published_date"] for p in papers if p["published_date"]})
        
        payload = {
            "count": len(papers),
            "crawled_date_min": crawled_dates[0] if crawled_dates else "",
            "crawled_date_max": crawled_dates[-1] if crawled_dates else "",
            "published_date_min": published_dates[0] if published_dates else "",
            "published_date_max": published_dates[-1] if published_dates else "",
            "papers": papers,
        }
        
        viewer_json.parent.mkdir(parents=True, exist_ok=True)
        viewer_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[INFO] Viewer JSON updated: {viewer_json} (count={len(papers)})")
        return True
        
    except Exception as e:
        print(f"[ERROR] Viewer JSON export failed: {e}")
        return False


def main():
    print("=" * 60)
    print("[START] LLM Paper Processing")
    print(f"Date: {date.today().isoformat()}")
    print(f"API Base: {API_BASE_URL}")
    print(f"Model: {MODEL_NAME}")
    print("=" * 60)
    
    if not API_KEY:
        print("[ERROR] LLM_API_KEY not configured!")
        print("Please set LLM_API_KEY environment variable or add to GitHub Secrets")
        return 1
    
    papers_dir = Path("papers")
    papers_dir.mkdir(parents=True, exist_ok=True)
    
    # 加载待处理论文
    pending_papers = load_pending_papers()
    if not pending_papers:
        print("[INFO] No pending papers to process")
        return 0
    
    print(f"[INFO] {len(pending_papers)} papers to process")
    
    # 处理每篇论文
    results = []
    for paper in pending_papers:
        result = process_paper(paper, papers_dir)
        results.append(result)
    
    # 更新 Excel
    print("\n[INFO] Updating Excel...")
    update_excel_with_results(results)
    
    # 导出 Viewer JSON
    print("\n[INFO] Exporting Viewer JSON...")
    export_viewer_json()
    
    # 清理待处理文件
    output_json = Path("new_papers.json")
    if output_json.exists():
        output_json.unlink()
    
    print("\n" + "=" * 60)
    print("[DONE] LLM processing completed")
    print("=" * 60)
    
    return 0


if __name__ == "__main__":
    exit(main())
