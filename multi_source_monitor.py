#!/usr/bin/env python3
"""
Multi-Source Paper Monitor (arXiv + Scopus)
双源文献监控 - GitHub Actions 版本
支持：
  - arXiv: 预印本（免费，无需 API key）
  - Scopus: 核心期刊（需注册 API key）
"""

import os
import sys
import json
import time
import random
import hashlib
import requests
from datetime import date, datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ==================== 配置 ====================
# arXiv 配置
ARXIV_API_URL = "https://export.arxiv.org/api/query"
ARXIV_REQUEST_INTERVAL = 6  # 秒

# Scopus 配置
SCOPUS_API_URL = "https://api.elsevier.com/content/search/scopus"
SCOPUS_API_KEY = os.environ.get("SCOPUS_API_KEY", "")  # 必需

# 通用配置
REQUEST_INTERVAL = 3  # 不同平台间请求间隔
MAX_RETRIES = 3
RETRY_DELAY = 10

PAPERS_DIR = Path("papers")
CRAWLED_IDS_FILE = Path("crawled_ids.txt")
EXCEL_FILE = Path("papers_record.xlsx")

# 飞书配置
FEISHU_APP_ID = os.environ.get("FEISHU_APP_ID", "")
FEISHU_APP_SECRET = os.environ.get("FEISHU_APP_SECRET", "")
FEISHU_CHAT_ID = os.environ.get("FEISHU_CHAT_ID", "")


def load_search_config():
    """加载分数据源的搜索配置"""
    config_file = Path("search_config.json")
    if not config_file.exists():
        # 默认配置
        return {
            "arxiv": "oceanography AND (deep learning OR neural network OR transformer OR AI OR machine learning)",
            "semantic_scholar": "oceanography AND machine learning OR deep learning OR AI",
            "scopus": "TITLE-ABS-KEY(oceanography AND (deep learning OR neural network OR AI OR machine learning))"
        }
    
    config = json.loads(config_file.read_text())
    return config


def generate_paper_id(source: str, identifier: str) -> str:
    """生成唯一论文 ID（用于去重）"""
    return f"{source}:{identifier}"


def search_arxiv_papers(ocean_keywords: str, ai_keywords: str, max_results: int = 20):
    """搜索 arXiv 论文（带重试）
    
    arXiv 使用 AND 连接两个查询条件
    """
    # 构建 arXiv 查询：(海洋学关键词) AND (AI 关键词)
    search_query = f"({ocean_keywords}) AND ({ai_keywords})"
    print(f"\n[arXiv] Searching: {search_query}")
    
    for attempt in range(MAX_RETRIES + 1):
        try:
            if attempt > 0:
                delay = RETRY_DELAY * (2 ** (attempt - 1)) + random.uniform(0.5, 2.0)
                print(f"[arXiv] Retrying in {delay:.1f}s...")
                time.sleep(delay)
            elif attempt == 0:
                time.sleep(REQUEST_INTERVAL)
            
            params = {
                "search_query": search_query,
                "max_results": max_results,
                "sortBy": "submittedDate",
                "sortOrder": "descending"
            }
            
            response = requests.get(
                ARXIV_API_URL,
                params=params,
                timeout=60,
                headers={"User-Agent": "Hermes-MultiSource-Monitor/1.0"}
            )
            
            if response.status_code == 429:
                print(f"[arXiv] Rate limited (429), attempt {attempt + 1}/{MAX_RETRIES}")
                if attempt < MAX_RETRIES:
                    continue
                raise Exception("arXiv rate limit exceeded")
            
            response.raise_for_status()
            
            import xml.etree.ElementTree as ET
            root = ET.fromstring(response.content)
            ns = {"atom": "http://www.w3.org/2005/Atom", "arxiv": "http://arxiv.org/schemas/atom"}
            
            papers = []
            for entry in root.findall("atom:entry", ns):
                title_elem = entry.find("atom:title", ns)
                title = title_elem.text.strip().replace("\n", " ") if title_elem is not None else "No title"
                
                id_elem = entry.find("atom:id", ns)
                arxiv_id = id_elem.text.split("/")[-1] if id_elem is not None else "unknown"
                
                summary_elem = entry.find("atom:summary", ns)
                abstract = summary_elem.text.strip().replace("\n", " ")[:500] if summary_elem is not None else "No abstract"
                
                published_elem = entry.find("atom:published", ns)
                published = published_elem.text[:10] if published_elem is not None else ""
                
                authors = []
                for author in entry.findall("atom:author", ns):
                    name_elem = author.find("atom:name", ns)
                    if name_elem is not None:
                        authors.append(name_elem.text)
                
                papers.append({
                    "id": generate_paper_id("arxiv", arxiv_id),
                    "source": "arXiv",
                    "arxiv_id": arxiv_id,
                    "title": title,
                    "abstract": abstract,
                    "published": published,
                    "authors": authors[:5],
                    "url": f"https://arxiv.org/abs/{arxiv_id}",
                })
            
            print(f"[arXiv] Found {len(papers)} papers")
            return papers
            
        except Exception as e:
            print(f"[arXiv] Error: {e}")
            if attempt >= MAX_RETRIES:
                return []
    
    return []


def search_semantic_scholar(keywords: str, max_results: int = 20):
    """搜索 Semantic Scholar 论文
    
    Semantic Scholar 使用自然语言查询：
    - 使用引号包裹短语，如 "physical oceanography"
    - 多个词用 OR 连接
    - 不支持复杂的布尔逻辑嵌套
    """
    print(f"\n[Semantic Scholar] Searching: {keywords}")
    
    headers = {"Content-Type": "application/json"}
    if SEMANTIC_SCHOLAR_API_KEY:
        headers["x-api-key"] = SEMANTIC_SCHOLAR_API_KEY
    
    # Semantic Scholar 直接使用自然语言查询
    # 支持 OR 连接多个关键词/短语
    query_text = keywords
    
    params = {
        "query": query_text,
        "fields": "title,abstract,publicationDate,authors,url,publicationTypes,journal,venue,year,openAccessPdf",
        "year": f"{(datetime.now().year - 2)}-",  # 最近 2-3 年
        "limit": max_results
    }
    
    for attempt in range(MAX_RETRIES + 1):
        try:
            if attempt > 0:
                time.sleep(RETRY_DELAY * (2 ** (attempt - 1)))
            elif attempt == 0:
                time.sleep(REQUEST_INTERVAL)
            
            response = requests.get(
                SEMANTIC_SCHOLAR_URL,
                params=params,
                headers=headers,
                timeout=60
            )
            
            if response.status_code == 429:
                print(f"[Semantic Scholar] Rate limited (429), attempt {attempt + 1}/{MAX_RETRIES}")
                if attempt < MAX_RETRIES:
                    continue
            
            response.raise_for_status()
            data = response.json()
            
            papers = []
            for paper in data.get("data", []):
                paper_id = paper.get("paperId", "unknown")
                title = paper.get("title", "No title")
                abstract = paper.get("abstract", "No abstract")[:500] if paper.get("abstract") else "No abstract"
                pub_date = paper.get("publicationDate", "")[:10] if paper.get("publicationDate") else ""
                
                authors = []
                for author in paper.get("authors", [])[:5]:
                    if author.get("name"):
                        authors.append(author["name"])
                
                pdf_url = paper.get("openAccessPdf", {}).get("url") if paper.get("openAccessPdf") else None
                url = pdf_url if pdf_url else paper.get("url", f"https://www.semanticscholar.org/paper/{paper_id}")
                
                papers.append({
                    "id": generate_paper_id("semantic_scholar", paper_id),
                    "source": "Semantic Scholar",
                    "paper_id": paper_id,
                    "title": title,
                    "abstract": abstract,
                    "published": pub_date,
                    "authors": authors,
                    "url": url,
                    "venue": paper.get("venue", paper.get("journal", {}).get("name", "")),
                })
                
                if len(papers) >= max_results:
                    break
            
            print(f"[Semantic Scholar] Found {len(papers)} papers")
            return papers
            
        except Exception as e:
            print(f"[Semantic Scholar] Error: {e}")
            if attempt >= MAX_RETRIES:
                return []
    
    return []


def search_scopus(keywords: str, max_results: int = 20):
    """搜索 Scopus 论文
    
    Scopus 使用专门的查询语法：
    - TITLE-ABS-KEY(): 搜索标题、摘要、关键词
    - AND/OR/NOT: 逻辑运算符
    - PUBYEAR >: 出版年份过滤
    """
    print(f"\n[Scopus] Searching: {keywords}")
    
    if not SCOPUS_API_KEY:
        print("[Scopus] API key not configured, skipping")
        return []
    
    headers = {
        "X-ELS-APIKey": SCOPUS_API_KEY,
        "Accept": "application/json"
    }
    
    # Scopus 查询已经是指定的语法，直接使用
    # 添加出版年份过滤（最近 3 年）
    scopus_query = f"{keywords} AND PUBYEAR > {datetime.now().year - 3}"
    
    params = {
        "query": scopus_query,
        "count": max_results,
        "start": 0,
        "sortBy": "prism_coverdate",
        "sortDirection": "desc",
        "view": "COMPLETE",  # 获取完整信息
    }
    
    for attempt in range(MAX_RETRIES + 1):
        try:
            if attempt > 0:
                time.sleep(RETRY_DELAY * (2 ** (attempt - 1)))
            elif attempt == 0:
                time.sleep(REQUEST_INTERVAL)
            
            response = requests.get(
                SCOPUS_API_URL,
                params=params,
                headers=headers,
                timeout=60
            )
            
            if response.status_code == 401 or response.status_code == 403:
                print(f"[Scopus] Authentication failed: {response.status_code}")
                print(f"[Scopus] Please check your API key at https://dev.elsevier.com")
                return []
            
            if response.status_code == 429:
                print(f"[Scopus] Rate limited (429), attempt {attempt + 1}/{MAX_RETRIES}")
                if attempt < MAX_RETRIES:
                    continue
            
            response.raise_for_status()
            data = response.json()
            
            papers = []
            
            # Scopus API 返回结构：search-results -> entry
            entries = data.get("search-results", {}).get("entry", [])
            
            for entry in entries:
                # 获取 DOI 或 Scopus ID
                doi = entry.get("prism_doi", "")
                scopus_id = entry.get("dc:identifier", "").replace("SCOPUS_ID:", "")
                paper_id = doi or scopus_id
                
                if not paper_id:
                    continue
                
                title = entry.get("dc:title", "No title")
                
                # 获取摘要
                abstract = entry.get("dc:description", "No abstract")
                if not abstract or abstract == "No abstract":
                    abstract = "No abstract available"
                abstract = abstract[:500]
                
                # 获取作者
                authors = []
                if "authkeywords" in entry:
                    # authkeywords 可能是一个列表或字典
                    auth_list = entry["authkeywords"]
                    if isinstance(auth_list, dict):
                        auth_list = auth_list.get("$text", "").split("|")
                    elif isinstance(auth_list, list):
                        auth_list = [item.get("$text", "") if isinstance(item, dict) else str(item) for item in auth_list]
                    else:
                        auth_list = str(auth_list).split("|")
                    
                    for i, kw in enumerate(auth_list[:5]):
                        authors.append(kw.strip())
                
                # 出版日期
                pub_date = entry.get("prism_coverdate", "")[:10] if entry.get("prism_coverdate") else ""
                
                # 链接
                url = entry.get("link", [{}])[0].get("@href", "") if entry.get("link") else ""
                if not url and doi:
                    url = f"https://doi.org/{doi}"
                elif not url and scopus_id:
                    url = f"https://www.scopus.com/record/display.uri?eid={scopus_id}"
                
                # 期刊信息
                journal = entry.get("prism_publicationName", "")
                
                papers.append({
                    "id": generate_paper_id("scopus", paper_id),
                    "source": "Scopus",
                    "doi": doi,
                    "scopus_id": scopus_id,
                    "title": title,
                    "abstract": abstract,
                    "published": pub_date,
                    "authors": authors,
                    "url": url,
                    "journal": journal,
                })
                
                if len(papers) >= max_results:
                    break
            
            print(f"[Scopus] Found {len(papers)} papers")
            return papers
            
        except json.JSONDecodeError as e:
            print(f"[Scopus] JSON parse error: {e}")
            print(f"[Scopus] Response: {response.text[:500]}")
            if attempt >= MAX_RETRIES:
                return []
        except Exception as e:
            print(f"[Scopus] Error: {e}")
            if attempt >= MAX_RETRIES:
                return []
    
    return []


def merge_and_deduplicate(all_papers_list: list) -> list:
    """合并多个来源的论文并去重"""
    seen_ids = set()
    merged = []
    
    for papers in all_papers_list:
        for paper in papers:
            if paper["id"] not in seen_ids:
                seen_ids.add(paper["id"])
                merged.append(paper)
    
    # 按出版日期排序
    merged.sort(key=lambda x: x.get("published", ""), reverse=True)
    
    return merged


def load_crawled_ids() -> set:
    """加载已爬取的论文 ID"""
    if CRAWLED_IDS_FILE.exists():
        return set(line.strip() for line in CRAWLED_IDS_FILE.read_text().splitlines() if line.strip())
    return set()


def save_crawled_ids(crawled_ids: set):
    """保存已爬取的论文 ID"""
    CRAWLED_IDS_FILE.write_text("\n".join(sorted(crawled_ids)))


def load_or_create_excel():
    """加载或创建 Excel 文件"""
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Papers" not in wb.sheetnames:
            wb.create_sheet("Papers")
        return wb
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Papers"
    headers = [
        "arxiv_id", "title", "authors", "affiliations",
        "published_date", "categories", "abstract", "summary_cn",
        "pdf_filename", "crawled_date", "notes", "source", "url"
    ]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 设置列宽
    col_widths = [15, 45, 28, 35, 12, 18, 70, 60, 20, 12, 25, 10, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    return wb


def append_to_excel(wb, paper: dict):
    """添加论文到 Excel"""
    ws = wb["Papers"]
    today = date.today().isoformat()
    
    # 从论文 ID 中提取 arxiv_id (格式：arxiv:1234.56789)
    arxiv_id = ""
    if paper.get("arxiv_id"):
        arxiv_id = paper["arxiv_id"]
    elif paper.get("id", "").startswith("arxiv:"):
        arxiv_id = paper["id"].split(":")[1]
    
    row = [
        arxiv_id,
        paper.get("title", ""),
        ", ".join(paper.get("authors", [])) if isinstance(paper.get("authors"), list) else paper.get("authors", ""),
        paper.get("affiliations", ""),
        paper.get("published", "")[:10] if paper.get("published") else "",
        paper.get("categories", ""),
        paper.get("abstract", "")[:500] if paper.get("abstract") else "",
        paper.get("summary_cn", ""),
        f"{arxiv_id}.pdf" if arxiv_id else "",
        today,
        "",  # notes
        paper.get("source", ""),
        paper.get("url", ""),
    ]
    ws.append(row)
    
    # 设置单元格格式
    last_row = ws.max_row
    for col in range(1, len(row) + 1):
        ws.cell(row=last_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    
    print(f"[INFO] Appended to Excel: {arxiv_id} - {paper.get('title', '')[:40]}...")


def save_excel(wb):
    """保存 Excel 文件"""
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)
    print(f"[INFO] Excel saved: {EXCEL_FILE}")


def export_viewer_json_from_excel():
    """从 papers_record.xlsx 导出 viewer 使用的 papers_data.json"""
    if not EXCEL_FILE.exists():
        print(f"[WARN] Excel not found, skip viewer export: {EXCEL_FILE}")
        return
    
    viewer_json = Path("viewer/papers_data.json")
    viewer_json.parent.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    if "Papers" not in wb.sheetnames:
        print("[WARN] Sheet 'Papers' not found, skip viewer export")
        return
    
    ws = wb["Papers"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        print("[WARN] Excel header missing, skip viewer export")
        return
    
    headers = [str(h) if h is not None else "" for h in header_row]
    index = {name: i for i, name in enumerate(headers)}
    
    required = [
        "arxiv_id", "title", "authors", "affiliations", "published_date",
        "categories", "abstract", "summary_cn", "pdf_filename", "crawled_date", "notes"
    ]
    missing = [c for c in required if c not in index]
    if missing:
        print(f"[WARN] Missing columns in Excel, skip viewer export: {missing}")
        return
    
    def norm(v):
        if v is None:
            return ""
        return str(v).replace("\n", " ").strip()
    
    def quality_key(p):
        return (
            1 if p.get("summary_cn") else 0,
            1 if p.get("affiliations") else 0,
            len(p.get("summary_cn", "")),
            len(p.get("affiliations", "")),
            len(p.get("abstract", "")),
            p.get("crawled_date", ""),
            p.get("published_date", ""),
        )
    
    papers_by_id = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        paper = {col: norm(row[index[col]]) for col in required}
        if not paper["arxiv_id"]:
            continue
        paper["pdf_url"] = f"https://arxiv.org/pdf/{paper['arxiv_id']}"
        arxiv_id = paper["arxiv_id"]
        old = papers_by_id.get(arxiv_id)
        if old is None or quality_key(paper) > quality_key(old):
            papers_by_id[arxiv_id] = paper
    
    papers = list(papers_by_id.values())
    papers.sort(key=lambda x: (x["crawled_date"], x["published_date"], x["arxiv_id"]), reverse=True)
    
    crawled_dates = sorted({p["crawled_date"] for p in papers if p["crawled_date"]})
    published_dates = sorted({p["published_date"] for p in papers if p["published_date"]})
    
    payload = {
        "count": len(papers),
        "crawled_date_min": crawled_dates[0] if crawled_dates else "",
        "crawled_date_max": crawled_dates[-1] if crawled_dates else "",
        "published_date_min": published_dates[0] if published_dates else "",
        "published_date_max": published_dates[-1] if published_dates else "",
        "papers": papers,
    }
    
    viewer_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[INFO] Viewer JSON updated: {viewer_json} (count={len(papers)})")


def get_feishu_access_token() -> str:
    """获取飞书访问令牌"""
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    payload = {"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET}
    
    response = requests.post(url, json=payload, timeout=30)
    result = response.json()
    
    if result.get("code") != 0:
        raise Exception(f"Failed to get Feishu token: {result}")
    
    return result["tenant_access_token"]


def send_feishu_message(message: str):
    """发送飞书消息"""
    if not FEISHU_CHAT_ID:
        print("[WARNING] FEISHU_CHAT_ID not set, skipping message send")
        return
    
    token = get_feishu_access_token()
    url = "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id"
    
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "receive_id": FEISHU_CHAT_ID,
        "msg_type": "text",
        "content": json.dumps({"text": message})
    }
    
    response = requests.post(url, headers=headers, json=payload, timeout=30)
    result = response.json()
    
    if result.get("code") != 0:
        print(f"[ERROR] Failed to send Feishu message: {result}")
    else:
        print(f"[INFO] Message sent to Feishu")


def build_message(new_papers: list, source_counts: dict) -> str:
    """构建飞书推送消息"""
    today = date.today().isoformat()
    
    message = f"🔬 **物理海洋学+AI 新论文推送**\n\n"
    message += f"📅 日期：{today}\n"
    message += f"📊 发现 {len(new_papers)} 篇新论文\n"
    message += f"   • arXiv: {source_counts.get('arXiv', 0)} 篇\n"
    message += f"   • Scopus: {source_counts.get('Scopus', 0)} 篇\n\n"
    
    message += "**最新论文列表**:\n\n"
    
    for i, paper in enumerate(new_papers[:10], 1):
        title = paper["title"][:60] + "..." if len(paper["title"]) > 60 else paper["title"]
        authors = ", ".join(paper.get("authors", [])[:3]) if paper.get("authors") else "Unknown"
        if len(paper.get("authors", [])) > 3:
            authors += " et al."
        
        source_emoji = {"arXiv": "📄", "Scopus": "📚"}.get(paper["source"], "📝")
        
        message += f"{i}. {source_emoji} [{title}]({paper['url']})\n"
        message += f"   作者：{authors} | {paper.get('published', 'N/A')} | {paper['source']}\n\n"
    
    if len(new_papers) > 10:
        message += f"\n... 还有 {len(new_papers) - 10} 篇论文，请查看 GitHub Pages 获取完整列表。\n"
    
    message += f"\n📌 搜索策略：oceanography + AI/ML/数据同化/PINN"
    
    return message


def main():
    print("=" * 70)
    print(f"[START] Multi-Source Paper Monitor (arXiv + Semantic Scholar + Scopus)")
    print(f"[INFO] Date: {date.today().isoformat()}")
    print("=" * 70)
    
    PAPERS_DIR.mkdir(parents=True, exist_ok=True)
    
    # 加载分数据源的搜索配置
    search_config = load_search_config()
    arxiv_ocean = search_config.get("arxiv", "")
    arxiv_ai = search_config.get("arxiv_ai", "")
    scopus_query = search_config.get("scopus", "")
    
    print(f"[INFO] arXiv 海洋学关键词：{arxiv_ocean}")
    print(f"[INFO] arXiv AI 关键词：{arxiv_ai}")
    print(f"[INFO] Scopus 查询：{scopus_query}")
    
    # 加载已爬取 ID
    crawled_ids = load_crawled_ids()
    print(f"[INFO] Loaded {len(crawled_ids)} crawled IDs")
    
    # 搜索两个来源：arXiv + Scopus
    all_papers = []
    source_counts = {}
    
    # 1. arXiv - 使用分离的海洋学和 AI 关键词
    arxiv_papers = search_arxiv_papers(arxiv_ocean, arxiv_ai, max_results=20)
    all_papers.append(arxiv_papers)
    source_counts["arXiv"] = len(arxiv_papers)
    
    # 2. Scopus - 使用 Scopus 专用语法
    scopus_papers = search_scopus(scopus_query, max_results=20)
    all_papers.append(scopus_papers)
    source_counts["Scopus"] = len(scopus_papers)
    
    # 合并去重
    all_unique_papers = merge_and_deduplicate(all_papers)
    print(f"\n[INFO] Total unique papers: {len(all_unique_papers)}")
    
    # 查重
    new_papers = [p for p in all_unique_papers if p["id"] not in crawled_ids]
    print(f"[INFO] New papers: {len(new_papers)}")
    
    if not new_papers:
        message = f"✅ 今日（{date.today().isoformat()}）未发现新的物理海洋学+AI 相关论文。\n\n"
        message += f"搜索来源：arXiv, Scopus\n"
        message += f"关键词：oceanography + AI/ML/数据同化/PINN"
        send_feishu_message(message)
        print("[INFO] No new papers. Notification sent.")
        return
    
    # 构建并发送消息
    message = build_message(new_papers, source_counts)
    send_feishu_message(message)
    
    # 更新已爬取 ID
    new_ids = {p["id"] for p in new_papers}
    crawled_ids.update(new_ids)
    save_crawled_ids(crawled_ids)
    print(f"[INFO] Updated crawled_ids.txt with {len(new_ids)} new IDs")
    
    # 更新 Excel 记录
    print("\\n[INFO] Updating Excel record...")
    wb = load_or_create_excel()
    for paper in new_papers:
        append_to_excel(wb, paper)
    save_excel(wb)
    
    # 导出 Viewer JSON
    print("\\n[INFO] Exporting Viewer JSON...")
    export_viewer_json_from_excel()
    
    print("\\n" + "=" * 70)
    print("[DONE] Multi-source monitor completed successfully")
    print("=" * 70)


if __name__ == "__main__":
    main()
